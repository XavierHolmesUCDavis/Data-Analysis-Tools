# -*- coding: utf-8 -*-
"""
Created on Wed Feb 12 05:24:27 2025

@author: Xavier Holmes

Script to make an excel sheet w/ protein accessions and their relative abundance for a group of samples
The group of samples will come from all excels in a directory 

"""

import time
import os
import pandas as pd
import numpy as np
import re
import openpyxl


experiment_name = input("Enter the name of the experiment: ")
time_1 = time.time()

# byonic_file_path = "Z:/Xavier/Second Year/November_2022/Cyt_1_R2/Adjpoint2/After_GoodQC_11_28/Cytokine_Experiment_Proteins"

byonic_file_path = 'Z:/Xavier/Second Year/November_2022/Cyt_1_R2/Adjpoint2/After_GoodQC_11_28/Cytokine_Experiment_Proteins/No Technical Replicates'

all_files = os.listdir(byonic_file_path)

combined_df = pd.DataFrame()  

for file in all_files:
    if file.endswith(".xlsx"):
        print("Processing file: ", file)
        workbook = openpyxl.load_workbook(byonic_file_path + "/" + file)
        ws = workbook['Proteins']
        protein_IDs = ws['B']
        protein_IDs = [str(cell.value) for cell in protein_IDs]
        protein_IDs = protein_IDs[1:]
        for i in range(len(protein_IDs)):
            protein_IDs[i] = re.search(r'\|(.*)\|', protein_IDs[i]).group(1)
        
        intensity = ws['F']
        intensity = [str(cell.value) for cell in intensity]
        intensity = intensity[1:]

        intensity = [0 if x == 'None' else x for x in intensity]
        intensity = [float(x) for x in intensity]
        intensity = np.array(intensity)
        total_intensity = np.sum(intensity)
        
        relative_abundance = []
        relative_abundance = (intensity / total_intensity)*100
        relative_abundance = np.array(relative_abundance)

        file_name = file.replace(".xlsx", "")
        file_name = re.sub(r'_raw.*', '', file_name)

        if combined_df.empty:
            combined_df = pd.DataFrame(relative_abundance, index=protein_IDs, columns=[file_name])
        else:
            combined_df = combined_df.join(pd.DataFrame(relative_abundance, index=protein_IDs, columns=[file_name]), how='outer')

time_2 = time.time()

combined_df = combined_df.fillna(0)
combined_df.to_excel(byonic_file_path + f"/Combined_Protein_Abundance_{experiment_name}.xlsx")

print("Grouping samples took ", (time_2 - time_1)/60, "minutes")


        
