# -*- coding: utf-8 -*-
"""
Created on Mon Apr  1 13:51:43 2024

@author: Xavier Holmes
"""

import openpyxl
import os 

byonic_files_folder = "Z:/Xavier/Second Year/November_2022/Cyt_1_R2/Adjpoint2/After_GoodQC_11_28/Cytokine_Experiment_Proteins"

for byonic_file in os.listdir(byonic_files_folder):
    if byonic_file.endswith(".xlsx"):
        print('Removing decoys from', byonic_file)
        try:
            wb = openpyxl.load_workbook(os.path.join(byonic_files_folder, byonic_file))
            ws = wb["Proteins"]
            rows_to_delete = []
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
                for cell in row:
                    if ">Reverse" in cell.value:
                        rows_to_delete.append(cell.row)
            for row in reversed(rows_to_delete):
                ws.delete_rows(row)
                
            wb.save(os.path.join(byonic_files_folder, byonic_file))
            wb.close()
        except TypeError:
            continue

